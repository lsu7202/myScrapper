"""
엑셀 매칭 로직
DB의 수집 데이터를 기존 엑셀 파일과 매칭

예:
  - DB의 (공시회사명, 대표자명)과
  - Excel의 (공시회사명, 대표자명) 비교
  - 일치하면 감사보고서, 제출인 정보 입력
"""

import os
import sys
import openpyxl
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db_models import SessionLocal, AuditReport

# ==================== 설정 ====================
EXCEL_FILE = os.path.getenv("EXCEL_FILE", "../기업개황.xlsx")

# ==================== 매칭 로직 ====================
class ExcelMatcher:
    """엑셀과 DB 데이터 매칭"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
        
        self.wb = openpyxl.load_workbook(excel_path)
        self.ws = self.wb.active
    
    def get_column_indices(self):
        """컬럼 인덱스 가져오기"""
        headers_row = self.ws[1]
        header_values = {cell.value: i+1 for i, cell in enumerate(headers_row)}
        
        required_cols = {
            '공시회사명': None,
            '대표자명': None,
        }
        
        for col_name in required_cols:
            if col_name not in header_values:
                raise ValueError(f"필수 컬럼을 찾을 수 없습니다: {col_name}")
            required_cols[col_name] = header_values[col_name]
        
        # 감사보고서, 제출인 컬럼 추가 (없으면 마지막 다음에 추가)
        if '감사보고서' not in header_values:
            report_col = len(header_values) + 1
            self.ws.cell(row=1, column=report_col, value='감사보고서')
        else:
            report_col = header_values['감사보고서']
        
        if '제출인' not in header_values:
            submitter_col = len(header_values) + 2
            self.ws.cell(row=1, column=submitter_col, value='제출인')
        else:
            submitter_col = header_values['제출인']
        
        return {
            'company': required_cols['공시회사명'],
            'ceo': required_cols['대표자명'],
            'report': report_col,
            'submitter': submitter_col
        }
    
    def normalize_string(self, text: str) -> str:
        """문자열 정규화 (비교용)"""
        if not text:
            return ""
        return ' '.join(str(text).strip().split())
    
    def get_db_data(self):
        """DB에서 데이터 조회 (정규화된 형태)"""
        db = SessionLocal()
        try:
            reports = db.query(AuditReport).all()
            
            # (company_name, ceo_name) → [report_data, ...]
            data_dict = {}
            
            for report in reports:
                company = self.normalize_string(report.company_name)
                ceo = self.normalize_string(report.ceo_name) if report.ceo_name else ""
                key = (company, ceo)
                
                if key not in data_dict:
                    data_dict[key] = {
                        'reports': [],
                        'submitter': report.submitter
                    }
                
                data_dict[key]['reports'].append(report.report_text)
            
            return data_dict
        
        finally:
            db.close()
    
    def match_and_update(self, db_data: dict) -> dict:
        """엑셀과 DB 데이터 매칭 및 업데이트"""
        cols = self.get_column_indices()
        
        stats = {
            'total_rows': 0,
            'matched_rows': 0,
            'failed_matches': []
        }
        
        # 데이터 행 순회 (1행은 헤더)
        for row in range(2, self.ws.max_row + 1):
            company_cell = self.ws.cell(row=row, column=cols['company'])
            ceo_cell = self.ws.cell(row=row, column=cols['ceo'])
            
            company_value = company_cell.value
            ceo_value = ceo_cell.value
            
            if not company_value:
                continue
            
            stats['total_rows'] += 1
            
            # 정규화
            company_norm = self.normalize_string(company_value)
            ceo_norm = self.normalize_string(ceo_value) if ceo_value else ""
            
            # DB에서 조회
            key = (company_norm, ceo_norm)
            
            if key in db_data:
                matched = db_data[key]
                
                # 감사보고서 기입 (여러 개면 / 로 구분)
                report_text = ' / '.join(matched['reports'])
                self.ws.cell(row=row, column=cols['report'], value=report_text)
                
                # 제출인 기입
                submitter = matched['submitter']
                if submitter:
                    self.ws.cell(row=row, column=cols['submitter'], value=submitter)
                
                stats['matched_rows'] += 1
                print(f"  ✓ Row {row}: {company_norm} / {ceo_norm}")
            
            else:
                # 매칭 실패 (로깅만)
                if ceo_norm:
                    # 대표자명이 있는데 매칭 안 됨
                    stats['failed_matches'].append({
                        'row': row,
                        'company': company_norm,
                        'ceo': ceo_norm,
                        'reason': 'CEO not found in DB'
                    })
                # 대표자명이 없는 경우는 무시
        
        return stats
    
    def save(self):
        """엑셀 저장"""
        self.wb.save(self.excel_path)
        print(f"  ✓ 엑셀 저장: {self.excel_path}")


def match_excel(excel_path: str = None) -> dict:
    """
    메인 매칭 함수
    
    Args:
        excel_path: 엑셀 파일 경로 (기본값: EXCEL_FILE)
    
    Returns:
        매칭 통계
    """
    if not excel_path:
        excel_path = EXCEL_FILE
    
    print(f"\n📊 엑셀 매칭 시작")
    print(f"  - 파일: {excel_path}\n")
    
    try:
        # 1. 엑셀 로드
        matcher = ExcelMatcher(excel_path)
        print(f"  ✓ 엑셀 로드 완료\n")
        
        # 2. DB 데이터 조회
        print(f"📥 DB에서 데이터 조회 중...")
        db_data = matcher.get_db_data()
        print(f"  ✓ {len(db_data)}개 데이터 조회\n")
        
        # 3. 매칭 및 업데이트
        print(f"🔍 매칭 중...")
        stats = matcher.match_and_update(db_data)
        print(f"  ✓ 매칭 완료\n")
        
        # 4. 저장
        print(f"💾 저장 중...")
        matcher.save()
        
        # 5. 통계 출력
        print(f"\n📋 매칭 결과:")
        print(f"  - 처리된 행: {stats['total_rows']}")
        print(f"  - 매칭된 행: {stats['matched_rows']}")
        print(f"  - 매칭률: {stats['matched_rows'] / stats['total_rows'] * 100:.1f}%" 
              if stats['total_rows'] > 0 else "  - 매칭률: N/A")
        
        if stats['failed_matches']:
            print(f"\n⚠️  매칭 실패 ({len(stats['failed_matches'])}개):")
            for fail in stats['failed_matches'][:5]:  # 처음 5개만 표시
                print(f"    - Row {fail['row']}: {fail['company']} / {fail['ceo']}")
                print(f"      (이유: {fail['reason']})")
            if len(stats['failed_matches']) > 5:
                print(f"    ... 외 {len(stats['failed_matches']) - 5}개")
        
        return stats
    
    except Exception as e:
        print(f"  ✗ 엑셀 매칭 실패: {str(e)}")
        raise


if __name__ == "__main__":
    # 테스트 실행
    match_excel()
